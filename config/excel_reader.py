# coding = utf-8
import openpyxl
import xlrd
from model.request_model import RequestModel


class ExcelReader:
    def reader(self, case_name):
        models_list = []
        wb = xlrd.open_workbook('./data/case.xlsx')
        names = wb.sheet_names()
        try:
            for sheet_name in names:
                if sheet_name == case_name:
                    sheet = wb.sheet_by_name(sheet_name)
                    for i in range(sheet.nrows):
                        if i == 0:
                            continue
                        smart_list = []
                        for j in range(sheet.ncols):
                            smart_list.append(sheet.cell_value(i, j))
                        model = RequestModel()
                        model.case_id = int(smart_list[0])
                        model.url_host = smart_list[1]
                        model.url_adr = smart_list[2]
                        model.module = smart_list[3]
                        model.title = smart_list[4]
                        model.is_run = smart_list[5]
                        model.method = smart_list[6]
                        model.req_param_type = smart_list[7]
                        model.pre_case_id = int(smart_list[8])
                        if smart_list[9] == '':
                            model.pre_fields = smart_list[9]
                        else:
                            model.pre_fields = eval(smart_list[9])
                        if smart_list[10] == '':
                            model.headers = smart_list[10]
                        else:
                            model.headers = eval(smart_list[10])
                        if smart_list[11] == '':
                            model.data = smart_list[11]
                        else:
                            model.data = eval(smart_list[11])
                        model.assert_type = smart_list[12]
                        model.assert_value = smart_list[13]
                        model.run_result = smart_list[14]
                        model.result_msg = smart_list[15]
                        model.response = smart_list[16]
                        model.update_time = smart_list[17]
                        models_list.append(model)
        except Exception as e:
            print(e)
        return models_list

    def writer(self, case_name, nrow, ncol, value):
        wb = openpyxl.load_workbook('./data/case.xlsx')
        sheet = wb[case_name]
        sheet.cell(nrow + 1, ncol).value = value
        wb.save('./data/case.xlsx')
        wb.close()

    def update_case(self, case_name, model, run_result, result_msg, response, update_time):
        wb = openpyxl.load_workbook('./data/case.xlsx')
        sheet = wb[case_name]
        sheet.cell(model.case_id + 1, 15).value = run_result
        sheet.cell(model.case_id + 1, 16).value = result_msg
        sheet.cell(model.case_id + 1, 17).value = response
        sheet.cell(model.case_id + 1, 18).value = update_time
        wb.save('./data/case.xlsx')
        wb.close()

if __name__ == '__main__':
    reader = ExcelReader()
    models = reader.reader('Sheet1')
    print(models)
    # reader.writer('Sheet1', 2, 15, 'bbb')
    reader.update_case('Sheet1',models[0],11,22,33,44)
